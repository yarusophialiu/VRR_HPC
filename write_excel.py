import pandas as pd
import re
import os
from utils import *
from openpyxl import Workbook, load_workbook



def create_df(fps_categories, values_dict):
    # new_row = ['bitrate'] + all_columns
    df = pd.DataFrame(columns=all_columns)
    # df = pd.DataFrame(columns=all_columns)
    df.loc[0] = [None] * len(all_columns)  # Initialize a new row with Nones

    # Insert the values into the DataFrame
    for fps in fps_categories:
        start_index = all_columns.index(360) + fps_categories.index(fps) * len(base_columns)
        end_index = start_index + len(base_columns)
        num_values = min(len(values_dict[fps]), len(base_columns))

        df.iloc[0, start_index:end_index] = values_dict[fps][:num_values] # row column  
    return df



def write_to_excel(df, excel_path):
    # if not os.path.exists(excel_path):
    mode = 'w'  if not os.path.exists(excel_path) else 'a'

    with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode, if_sheet_exists='overlay') as writer:

        if sheet_name in writer.book.sheetnames:
        # if mode == 'a' and sheet_name in writer.book.sheetnames:
            # Load existing Sheet X
            startrow = writer.book[sheet_name].max_row
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
        else:
            df.to_excel(writer, sheet_name=sheet_name, index=False)


# download cvvdp results from HPC
# process results by running this file
# write to excel using write_excel.py
# plot using plot_cvvdp.py

# close excel file to write data to it
if __name__ == "__main__":
    # BASE_DIR = "bistro"
    SCENE = "bistro"
    CLEANED_DIR = "cleaned"
    WRITE_EXCEL = True


    cleaned_scene_path = f'{CLEANED_DIR}/{SCENE}'

    excel_dir = r'C:\Users\15142\Projects\VRR\VRR_Plot_HPC'
    excel_path = f'{excel_dir}/data-08-09/bistro.xlsx'
    jobid_list = [i for i in range(1, 2)]

    for jobid in jobid_list:
        path, seg, speed = mapIdToPath(jobid-1)
        print(f'path, seg, speed {path, seg, speed}\n')

        file_path = f'{cleaned_scene_path}/{SCENE}_{jobid}.txt'
        sheet_name = f'path{path}_seg{seg}_{speed}'  # Specify your desired sheet name here

        if WRITE_EXCEL:
            try:
                wb = load_workbook(excel_path)
            except FileNotFoundError:
                wb = Workbook()
                print(f'Created a new workbook: {excel_path}')

            if sheet_name not in wb.sheetnames:
                wb.create_sheet(title=sheet_name)
                print(f'Sheet "{sheet_name}" created.')                    
            wb.save(excel_path)

        with open(file_path, 'r') as file:
            content = file.read()
        
        pattern = r"========================= bitrate (\d+) =========================(.*?)((?========================== bitrate)|$)"
        matches = re.findall(pattern, content, re.DOTALL)
        # print(f'matches \n{matches}')

        bitrate_data = {}
        for match in matches:
            bitrate = int(match[0])
            data = match[1].strip()
            fps_sections = re.split(r'========================= fps(\d+) =========================', data)
            fps_data = {}
            for i in range(1, len(fps_sections), 2):
                fps = int(fps_sections[i])
                cvvdp_values = re.findall(r'cvvdp=([\d\.]+)', fps_sections[i+1])
                fps_data[fps] = list(map(float, cvvdp_values))

            bitrate_data[bitrate] = fps_data # fps_data is like values_dict

            fps_categories = list(fps_data.keys())
            print(f'bitrate {bitrate}, fps data \n {fps_data}')
            fps_data = {key: value[1:] + [value[0]] for key, value in fps_data.items()}
            print(f'bitrate {bitrate}, fps data \n {fps_data}\n\n\n')

            # print(f'fps_categories {list(fps_data.keys())}')
            fps_categories = sorted(fps_categories)
            # print(f'fps_categories {fps_categories}')

            if WRITE_EXCEL:
                base_columns = [360, 480, 720, 864, 1080]
                all_columns = base_columns * 10
                df = create_df(fps_categories, fps_data)
                # excel_df = pd.read_excel(excel_path, sheet_name=sheet_name)
                df.insert(0, 'bitrate', bitrate)

                write_to_excel(df, excel_path) # 360, 480, 720, 864, 1080